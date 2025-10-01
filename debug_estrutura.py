"""
Script de debug para ver a estrutura real dos dados
"""
import requests
import json
from openpyxl import load_workbook

print("="*70)
print("🔍 DEBUG - Estrutura dos Dados")
print("="*70)
print()

# 1. Ver resposta da API
print("1️⃣ Fazendo requisição para API...")
payload = {
    "municipio": "TIMON",
    "periodo_inicio": "2015-01-01",
    "periodo_fim": "2024-12-31",
    "ajuizamento": "2020-01-01",
    "citacao": "2020-06-01",
    "honorarios_perc": 20.0,
    "honorarios_fixo": 0.0,
    "desagio_principal": 0.0,
    "desagio_honorarios": 0.0,
    "correcao_ate": "2024-12-31"
}

response = requests.post("http://localhost:8000/api/calcular", json=payload, timeout=180)
data = response.json()

print("✅ Resposta recebida")
print()
print("📊 Chaves dos outputs:")
for key in data['outputs'].keys():
    print(f"  - '{key}'")
print()

# Mostrar primeiro cenário
first_key = list(data['outputs'].keys())[0]
print(f"📄 Exemplo de cenário ('{first_key}'):")
print(json.dumps(data['outputs'][first_key], indent=2))
print()

# 2. Ver estrutura do Excel
print("2️⃣ Lendo planilha Excel...")
wb = load_workbook("data/timon_01-2025.xlsx", data_only=True)
ws = wb['RESUMO']

print("✅ Planilha lida")
print()
print("📊 Valores nas linhas 23-50 (coluna D):")
for linha in range(23, 51):
    valor_d = ws[f'D{linha}'].value
    valor_e = ws[f'E{linha}'].value
    valor_f = ws[f'F{linha}'].value
    if valor_d is not None:
        print(f"  Linha {linha}: D={valor_d} | E={valor_e} | F={valor_f}")
        
wb.close()
