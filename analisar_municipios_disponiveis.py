"""
Análise do Excel: Descobrir quais municípios estão disponíveis
"""
from pathlib import Path
from openpyxl import load_workbook

excel_path = Path("data/timon_01-2025.xlsx")

print("=" * 80)
print("ANÁLISE DO EXCEL: Municípios Disponíveis")
print("=" * 80)

wb = load_workbook(excel_path, data_only=True)

# Verificar aba RESUMO
if "RESUMO" in wb.sheetnames:
    ws = wb["RESUMO"]
    print("\n📋 Aba RESUMO encontrada")
    print(f"   B6 (Município): {ws['B6'].value}")
    print(f"   E6 (Período início): {ws['E6'].value}")
    print(f"   F6 (Período fim): {ws['F6'].value}")

# Listar todas as abas
print(f"\n📂 Total de abas: {len(wb.sheetnames)}")
print("\nAbas encontradas:")

# Separar por tipo
abas_calculos = []
abas_dados = []
abas_outras = []

for sheet_name in wb.sheetnames:
    if any(x in sheet_name.upper() for x in ['NT7', 'NT6', 'NT36', 'JASA']):
        abas_calculos.append(sheet_name)
    elif any(x in sheet_name.upper() for x in ['IPCA', 'SELIC', 'TR']):
        abas_dados.append(sheet_name)
    else:
        abas_outras.append(sheet_name)

print(f"\n📊 Abas de Cálculo ({len(abas_calculos)}):")
for aba in abas_calculos[:10]:  # Limitar a 10
    print(f"   - {aba}")
if len(abas_calculos) > 10:
    print(f"   ... e mais {len(abas_calculos) - 10}")

print(f"\n📈 Abas de Dados ({len(abas_dados)}):")
for aba in abas_dados:
    print(f"   - {aba}")

print(f"\n📄 Outras Abas ({len(abas_outras)}):")
for aba in abas_outras:
    print(f"   - {aba}")

# Verificar se há dados de outros municípios
print("\n" + "=" * 80)
print("VERIFICANDO DADOS PRÉ-CALCULADOS")
print("=" * 80)

# Ler alguns valores da linha 23 (primeiro cenário)
ws_resumo = wb["RESUMO"]
d23 = ws_resumo['D23'].value
d24 = ws_resumo['D24'].value

print(f"\nD23 (Principal BRUTO): {d23}")
print(f"D24 (Principal com deságio): {d24}")

if d23 is not None and d23 > 0:
    print("\n✅ Excel tem valores pré-calculados")
    print(f"   Valor: R$ {d23:,.2f}")
else:
    print("\n⚠️ Excel NÃO tem valores pré-calculados (zeros ou None)")

wb.close()

print("\n" + "=" * 80)
print("CONCLUSÃO")
print("=" * 80)
print(f"""
O Excel template contém:
- 1 município base: TIMON
- {len(abas_calculos)} cenários de cálculo
- {len(abas_dados)} abas de dados (taxas)
- Valores pré-calculados: {'SIM' if d23 and d23 > 0 else 'NÃO'}

Para testes válidos, devemos:
1. Usar TIMON como município
2. Usar períodos compatíveis com os dados do Excel
3. Comparar valores calculados com os valores do template
""")
print("=" * 80)
