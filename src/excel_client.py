"""
Cliente para manipular Excel local (openpyxl)
"""
import logging
from typing import Any, List
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
import config

logger = logging.getLogger(__name__)


class ExcelClient:
    """Cliente para interagir com arquivo Excel local"""
    
    def __init__(self):
        self.workbook = None
        self.excel_path = config.EXCEL_FULL_PATH
        
    def open_workbook(self, data_only=False):
        """
        Abre a planilha Excel
        
        Args:
            data_only: Se True, lê valores calculados. Se False, lê fórmulas (para escrita)
        """
        try:
            logger.info(f"Abrindo planilha: {self.excel_path} (data_only={data_only})")
            self.workbook = load_workbook(self.excel_path, data_only=data_only)
            logger.info("Planilha aberta com sucesso")
        except FileNotFoundError:
            logger.error(f"Arquivo Excel não encontrado: {self.excel_path}")
            raise FileNotFoundError(
                f"Planilha não encontrada: {self.excel_path}\n"
                f"Certifique-se de que o arquivo está na pasta do projeto."
            )
        except Exception as e:
            logger.error(f"Erro ao abrir planilha: {e}")
            raise
    
    def close_workbook(self):
        """Fecha a planilha (sem salvar)"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            logger.info("Planilha fechada")
    
    def get_worksheet(self, sheet_name: str) -> Worksheet:
        """Obtém uma aba da planilha"""
        if not self.workbook:
            raise ValueError("Planilha não está aberta")
        
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Aba '{sheet_name}' não encontrada. Abas disponíveis: {self.workbook.sheetnames}")
        
        return self.workbook[sheet_name]
    
    def write_cell(self, worksheet: str, address: str, value: Any):
        """
        Escreve um valor em uma célula
        
        Args:
            worksheet: Nome da aba
            address: Endereço da célula (ex: 'B6')
            value: Valor a escrever
        """
        ws = self.get_worksheet(worksheet)
        
        # Converter data para formato reconhecível
        if isinstance(value, str) and len(value) == 10 and value.count('-') == 2:
            # Formato YYYY-MM-DD, converter para datetime
            try:
                value = datetime.strptime(value, "%Y-%m-%d")
            except:
                pass
        
        ws[address] = value
        logger.debug(f"Escrito {value} em {worksheet}!{address}")
    
    def read_cell(self, worksheet: str, address: str) -> Any:
        """
        Lê valor de uma célula
        
        Args:
            worksheet: Nome da aba
            address: Endereço da célula (ex: 'D24')
            
        Returns:
            Valor da célula
        """
        ws = self.get_worksheet(worksheet)
        value = ws[address].value
        logger.debug(f"Lido {worksheet}!{address} = {value}")
        return value
    
    def read_range(self, worksheet: str, address: str) -> List[List[Any]]:
        """
        Lê valores de um range
        
        Args:
            worksheet: Nome da aba
            address: Endereço do range (ex: 'D24:F69')
            
        Returns:
            Matriz de valores
        """
        ws = self.get_worksheet(worksheet)
        
        # Parse do range (ex: "D24:F69")
        cells = ws[address]
        
        values = []
        for row in cells:
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            values.append(row_values)
        
        logger.debug(f"Lido range {worksheet}!{address}: {len(values)} linhas")
        
        return values
    
    def calculate(self):
        """
        Excel local calcula automaticamente ao abrir/modificar
        Não é necessário chamar calculate explicitamente
        """
        logger.info("Excel local: cálculos são automáticos")
        # Para forçar recálculo, podemos reabrir a planilha
        # Mas geralmente não é necessário
    
    def save_workbook(self):
        """Salva a planilha (use com cuidado!)"""
        if self.workbook:
            logger.warning("Salvando alterações na planilha...")
            self.workbook.save(self.excel_path)
            logger.info("Planilha salva")
    
    def get_workbook_version(self) -> str:
        """Obtém informação sobre a planilha"""
        try:
            import hashlib
            with open(self.excel_path, 'rb') as f:
                file_hash = hashlib.md5(f.read()).hexdigest()
            return f"md5:{file_hash[:8]}"
        except:
            return "local-file"
