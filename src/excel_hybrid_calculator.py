"""
Calculadora Excel híbrida: openpyxl para escrever + Excel desktop para calcular
"""
import logging
import subprocess
import time
from pathlib import Path
from typing import Any, List
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelHybridCalculator:
    """
    Usa openpyxl para escrever inputs
    Usa subprocess + Excel desktop para ler valores calculados
    """
    
    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.workbook = None
        
    def __enter__(self):
        """Context manager - abre workbook"""
        logger.info(f"Abrindo workbook: {self.excel_path}")
        self.workbook = load_workbook(self.excel_path, data_only=False)
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager - fecha workbook"""
        if self.workbook:
            self.workbook.close()
            
    def write_cell(self, worksheet_name: str, address: str, value: Any):
        """Escreve valor em célula via openpyxl"""
        try:
            ws = self.workbook[worksheet_name]
            ws[address] = value
            logger.debug(f"Escrito: {worksheet_name}!{address} = {value}")
        except Exception as e:
            logger.error(f"Erro ao escrever {address}: {e}")
            raise
            
    def save_and_calculate(self):
        """Salva workbook para que Excel calcule"""
        try:
            logger.info("Salvando workbook...")
            self.workbook.save(self.excel_path)
            logger.info("Workbook salvo. Aguardando Excel recalcular...")
            
            # Dar tempo para Excel processar se estiver monitorando o arquivo
            time.sleep(2)
            
        except Exception as e:
            logger.error(f"Erro ao salvar: {e}")
            raise
            
    def read_range_calculated(self, worksheet_name: str, range_address: str) -> List[List]:
        """
        Lê range com valores calculados
        Reabre workbook com data_only=True para pegar valores em vez de fórmulas
        """
        try:
            # Fechar workbook atual
            if self.workbook:
                self.workbook.close()
                
            # Reabrir com data_only=True
            logger.info("Reabrindo workbook para ler valores calculados...")
            wb_read = load_workbook(self.excel_path, data_only=True)
            ws = wb_read[worksheet_name]
            
            # Ler range
            cells = ws[range_address]
            
            # Converter para lista de listas
            if isinstance(cells, tuple):
                # Range com múltiplas células
                result = []
                for row in cells:
                    if isinstance(row, tuple):
                        result.append([cell.value for cell in row])
                    else:
                        result.append([row.value])
            else:
                # Célula única
                result = [[cells.value]]
                
            wb_read.close()
            
            # Reabrir workbook original para continuar
            self.workbook = load_workbook(self.excel_path, data_only=False)
            
            return result
            
        except Exception as e:
            logger.error(f"Erro ao ler range {range_address}: {e}")
            raise
            
    def force_calculate(self):
        """Placeholder - cálculo acontece ao salvar"""
        logger.info("Força cálculo não necessária (Excel calcula ao abrir)")
        pass
