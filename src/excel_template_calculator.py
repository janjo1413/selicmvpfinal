"""
Calculadora Excel com template protegido
Usa arquivo template (com valores calculados) como base
Cria cópias temporárias para cada cálculo
"""
import logging
import shutil
import tempfile
from pathlib import Path
from typing import Any, List
from openpyxl import load_workbook
from excel_recalculator import ExcelRecalculator

logger = logging.getLogger(__name__)


class ExcelTemplateCalculator:
    """
    Mantém arquivo template intacto (com valores pré-calculados)
    Cria cópia temporária para cada cálculo
    """
    
    def __init__(self, template_path: Path):
        self.template_path = template_path
        self.temp_path = None
        self.workbook = None
        
    def __enter__(self):
        """Context manager - cria cópia temporária"""
        # Criar arquivo temporário
        temp_dir = Path(tempfile.gettempdir())
        self.temp_path = temp_dir / f"calc_{self.template_path.name}"
        
        logger.info(f"Criando cópia temporária: {self.temp_path}")
        shutil.copy2(self.template_path, self.temp_path)
        
        logger.info(f"Abrindo cópia temporária...")
        self.workbook = load_workbook(self.temp_path, data_only=False)
        
        return self
        
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager - fecha e deleta cópia"""
        if self.workbook:
            self.workbook.close()
            
        # Deletar arquivo temporário
        if self.temp_path and self.temp_path.exists():
            try:
                self.temp_path.unlink()
                logger.info(f"Cópia temporária deletada: {self.temp_path}")
            except Exception as e:
                logger.warning(f"Erro ao deletar temp: {e}")
            
    def write_cell(self, worksheet_name: str, address: str, value: Any):
        """Escreve valor em célula"""
        try:
            ws = self.workbook[worksheet_name]
            ws[address] = value
            logger.debug(f"Escrito: {worksheet_name}!{address} = {value}")
        except Exception as e:
            logger.error(f"Erro ao escrever {address}: {e}")
            raise
            
    def save_workbook(self):
        """Salva workbook temporário (necessário para recálculo)"""
        try:
            self.workbook.save(self.temp_path)
            logger.debug(f"Workbook salvo: {self.temp_path}")
        except Exception as e:
            logger.error(f"Erro ao salvar workbook: {e}")
            raise
    
    def recalculate_workbook(self) -> bool:
        """
        Força recálculo do Excel usando win32com
        Deve ser chamado após escrever valores e antes de ler resultados
        
        Returns:
            True se recalculou com sucesso, False caso contrário
        """
        if not ExcelRecalculator.is_available():
            logger.warning("⚠️ win32com não disponível - pulando recálculo Excel")
            return False
        
        try:
            # Fechar workbook openpyxl e aguardar
            if self.workbook:
                self.workbook.close()
                self.workbook = None
            
            # Aguardar para garantir que arquivo foi liberado
            import time
            time.sleep(1.0)
            
            # Recalcular usando Excel COM
            logger.info("🔄 Recalculando Excel via win32com...")
            success = ExcelRecalculator.recalculate_workbook(self.temp_path, visible=False)
            
            if success:
                logger.info("✅ Excel recalculado com sucesso")
            else:
                logger.error("❌ Falha ao recalcular Excel")
            
            # Aguardar antes de reabrir
            time.sleep(1.0)
            
            # Reabrir workbook openpyxl
            self.workbook = load_workbook(self.temp_path, data_only=False)
            
            return success
            
        except Exception as e:
            logger.error(f"Erro ao recalcular workbook: {e}")
            # Tentar reabrir workbook mesmo em caso de erro
            if not self.workbook:
                try:
                    import time
                    time.sleep(1.0)
                    self.workbook = load_workbook(self.temp_path, data_only=False)
                except:
                    pass
            return False
            
    def read_range_calculated(self, worksheet_name: str, range_address: str) -> List[List]:
        """
        Lê range com valores calculados
        Fecha e reabre com data_only=True
        """
        try:
            # Fechar workbook atual
            if self.workbook:
                self.workbook.close()
                
            # Reabrir com data_only=True para ler valores
            logger.debug(f"Reabrindo temp com data_only=True para {range_address}")
            wb_read = load_workbook(self.temp_path, data_only=True)
            ws = wb_read[worksheet_name]
            
            # Ler range
            cells = ws[range_address]
            
            # Converter para lista de listas
            if isinstance(cells, tuple):
                # Range com múltiplas células
                result = []
                for row in cells:
                    if isinstance(row, tuple):
                        result.append([cell.value for cell in row])
                    else:
                        result.append([row.value])
            else:
                # Célula única
                result = [[cells.value]]
                
            wb_read.close()
            
            # Reabrir workbook para continuar (caso necessário)
            self.workbook = load_workbook(self.temp_path, data_only=False)
            
            return result
            
        except Exception as e:
            logger.error(f"Erro ao ler range {range_address}: {e}")
            raise
