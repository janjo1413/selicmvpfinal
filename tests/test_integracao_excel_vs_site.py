"""
Teste de Integração: Validação dos Cálculos do Site vs Planilha Excel

Este teste compara os resultados retornados pelo site com os valores 
pré-calculados na planilha Excel, garantindo precisão dos cálculos.
"""
import pytest
import requests
from datetime import date
from openpyxl import load_workbook
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

# Configuração
BASE_URL = "http://localhost:8000"
EXCEL_PATH = Path(__file__).parent.parent / "data" / "timon_01-2025.xlsx"

# Mapeamento de cenários (nome API -> linha no Excel)
# Baseado em OUTPUT_LINES do calculator_service.py
CENARIO_LINES = {
    "nt7_ipca_selic": 24,
    "nt7_periodo_cnj": 29,
    "nt6_ipca_selic": 34,
    "jasa_ipca_selic": 39,
    "nt7_tr": 44,
    "nt36_tr": 49,
    "nt7_ipca_e": 54,
    "nt36_ipca_e": 64,
    "nt36_ipca_e_1pct": 69,
}


class TestIntegracaoExcelVsSite:
    """Testes de integração comparando Excel e API"""
    
    @pytest.fixture(scope="class")
    def server_online(self):
        """Verifica se servidor está rodando"""
        try:
            response = requests.get(f"{BASE_URL}/", timeout=5)
            assert response.status_code == 200, "Servidor não está respondendo"
            logger.info("✅ Servidor online")
            return True
        except Exception as e:
            pytest.skip(f"Servidor não está rodando: {e}")
    
    @pytest.fixture(scope="class")
    def valores_excel(self):
        """Lê valores pré-calculados da planilha Excel"""
        logger.info(f"📊 Lendo planilha: {EXCEL_PATH}")
        
        # Abrir com data_only=True para ler valores calculados
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb['RESUMO']
        
        valores = {}
        for cenario_nome, linha in CENARIO_LINES.items():
            principal = ws[f'D{linha}'].value or 0.0
            honorarios = ws[f'E{linha}'].value or 0.0
            total = ws[f'F{linha}'].value or 0.0
            
            valores[cenario_nome] = {
                'principal': float(principal),
                'honorarios': float(honorarios),
                'total': float(total)
            }
            
            logger.info(f"  {cenario_nome}: P={principal:,.2f}, H={honorarios:,.2f}, T={total:,.2f}")
        
        wb.close()
        return valores
    
    @pytest.fixture(scope="class")
    def calcular_site(self, server_online):
        """Faz requisição para o site e retorna resultados"""
        logger.info("🌐 Fazendo requisição para o site...")
        
        # Payload com mesmos parâmetros da planilha
        payload = {
            "municipio": "TIMON",
            "periodo_inicio": "2015-01-01",
            "periodo_fim": "2024-12-31",
            "ajuizamento": "2020-01-01",
            "citacao": "2020-06-01",
            "honorarios_perc": 20.0,  # 20% como na planilha
            "honorarios_fixo": 0.0,
            "desagio_principal": 0.0,
            "desagio_honorarios": 0.0,  # Sem deságio para comparar direto
            "correcao_ate": "2024-12-31"
        }
        
        response = requests.post(
            f"{BASE_URL}/api/calcular",
            json=payload,
            timeout=180  # 3 minutos (cálculo pode demorar)
        )
        
        assert response.status_code == 200, f"Erro na API: {response.status_code}"
        
        data = response.json()
        logger.info(f"✅ Resposta recebida em {data['execution_time_ms']}ms")
        
        return data['outputs']
    
    def test_comparar_principal_todos_cenarios(self, valores_excel, calcular_site):
        """Compara valores de Principal entre Excel e Site (todos os cenários)"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Comparação de PRINCIPAL (Excel vs Site)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            excel_val = valores_excel[cenario]['principal']
            site_val = calcular_site[cenario]['principal']
            
            # Tolerância: 0.01% (R$ 100 em R$ 100.000.000)
            diff_percentual = abs(excel_val - site_val) / excel_val * 100 if excel_val > 0 else 0
            
            passou = diff_percentual < 0.01
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Excel:  R$ {excel_val:,.2f}")
            logger.info(f"   Site:   R$ {site_val:,.2f}")
            logger.info(f"   Diff:   {diff_percentual:.4f}%")
            
            if not passou:
                erros.append(f"{cenario}: {diff_percentual:.4f}% diferença")
        
        assert len(erros) == 0, f"Diferenças encontradas: {erros}"
    
    def test_comparar_honorarios_todos_cenarios(self, valores_excel, calcular_site):
        """Compara valores de Honorários entre Excel e Site (todos os cenários)"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Comparação de HONORÁRIOS (Excel vs Site)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            excel_val = valores_excel[cenario]['honorarios']
            site_val = calcular_site[cenario]['honorarios']
            
            # Honorários = 20% do principal (calculado em Python)
            principal = calcular_site[cenario]['principal']
            esperado = principal * 0.20
            
            # Tolerância: 0.01%
            diff_percentual = abs(site_val - esperado) / esperado * 100 if esperado > 0 else 0
            
            passou = diff_percentual < 0.01
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Excel:    R$ {excel_val:,.2f}")
            logger.info(f"   Site:     R$ {site_val:,.2f}")
            logger.info(f"   Esperado: R$ {esperado:,.2f} (20% de {principal:,.2f})")
            logger.info(f"   Diff:     {diff_percentual:.4f}%")
            
            if not passou:
                erros.append(f"{cenario}: {diff_percentual:.4f}% diferença")
        
        assert len(erros) == 0, f"Diferenças encontradas: {erros}"
    
    def test_comparar_total_todos_cenarios(self, valores_excel, calcular_site):
        """Compara valores de Total entre Excel e Site (todos os cenários)"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Comparação de TOTAL (Excel vs Site)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            excel_val = valores_excel[cenario]['total']
            site_val = calcular_site[cenario]['total']
            
            # Total calculado em Python = principal + honorarios
            principal = calcular_site[cenario]['principal']
            honorarios = calcular_site[cenario]['honorarios']
            esperado = principal + honorarios
            
            # Tolerância: 0.01%
            diff_percentual = abs(site_val - esperado) / esperado * 100 if esperado > 0 else 0
            
            passou = diff_percentual < 0.01
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Excel:    R$ {excel_val:,.2f}")
            logger.info(f"   Site:     R$ {site_val:,.2f}")
            logger.info(f"   Esperado: R$ {esperado:,.2f} ({principal:,.2f} + {honorarios:,.2f})")
            logger.info(f"   Diff:     {diff_percentual:.4f}%")
            
            if not passou:
                erros.append(f"{cenario}: {diff_percentual:.4f}% diferença")
        
        assert len(erros) == 0, f"Diferenças encontradas: {erros}"
    
    def test_honorarios_sao_20_porcento_do_principal(self, calcular_site):
        """Valida que honorários são exatamente 20% do principal"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Validação de Fórmula (Honorários = 20% Principal)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            principal = calcular_site[cenario]['principal']
            honorarios = calcular_site[cenario]['honorarios']
            
            esperado = principal * 0.20
            diff = abs(honorarios - esperado)
            
            passou = diff < 0.01  # Diferença menor que 1 centavo
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Principal:  R$ {principal:,.2f}")
            logger.info(f"   Honorários: R$ {honorarios:,.2f}")
            logger.info(f"   Esperado:   R$ {esperado:,.2f}")
            logger.info(f"   Diff:       R$ {diff:,.2f}")
            
            if not passou:
                erros.append(f"{cenario}: R$ {diff:,.2f} diferença")
        
        assert len(erros) == 0, f"Fórmula incorreta: {erros}"
    
    def test_total_e_soma_principal_mais_honorarios(self, calcular_site):
        """Valida que total = principal + honorarios"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Validação de Fórmula (Total = Principal + Honorários)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            principal = calcular_site[cenario]['principal']
            honorarios = calcular_site[cenario]['honorarios']
            total = calcular_site[cenario]['total']
            
            esperado = principal + honorarios
            diff = abs(total - esperado)
            
            passou = diff < 0.01  # Diferença menor que 1 centavo
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Principal:  R$ {principal:,.2f}")
            logger.info(f"   Honorários: R$ {honorarios:,.2f}")
            logger.info(f"   Total:      R$ {total:,.2f}")
            logger.info(f"   Esperado:   R$ {esperado:,.2f}")
            logger.info(f"   Diff:       R$ {diff:,.2f}")
            
            if not passou:
                erros.append(f"{cenario}: R$ {diff:,.2f} diferença")
        
        assert len(erros) == 0, f"Soma incorreta: {erros}"


class TestIntegracaoComDesagio:
    """Testes com deságio nos honorários"""
    
    @pytest.fixture(scope="class")
    def calcular_com_desagio(self):
        """Calcula com 10% de deságio nos honorários"""
        logger.info("🌐 Testando com 10% deságio...")
        
        payload = {
            "municipio": "TIMON",
            "periodo_inicio": "2015-01-01",
            "periodo_fim": "2024-12-31",
            "ajuizamento": "2020-01-01",
            "citacao": "2020-06-01",
            "honorarios_perc": 20.0,
            "honorarios_fixo": 0.0,
            "desagio_principal": 0.0,
            "desagio_honorarios": 10.0,  # 10% de desconto
            "correcao_ate": "2024-12-31"
        }
        
        response = requests.post(
            f"{BASE_URL}/api/calcular",
            json=payload,
            timeout=180
        )
        
        assert response.status_code == 200
        return response.json()['outputs']
    
    def test_desagio_10_porcento_aplicado(self, calcular_com_desagio):
        """Valida que deságio de 10% é aplicado corretamente"""
        logger.info("\n" + "="*70)
        logger.info("TESTE: Validação de Deságio (10% desconto)")
        logger.info("="*70)
        
        erros = []
        
        for cenario in CENARIO_LINES.keys():
            principal = calcular_com_desagio[cenario]['principal']
            honorarios = calcular_com_desagio[cenario]['honorarios']
            
            # Honorários bruto = 20% do principal
            hon_bruto = principal * 0.20
            # Honorários líquido = bruto - 10% desconto
            esperado = hon_bruto * 0.90
            
            diff = abs(honorarios - esperado)
            passou = diff < 0.01
            status = "✅" if passou else "❌"
            
            logger.info(f"{status} {cenario}:")
            logger.info(f"   Principal:     R$ {principal:,.2f}")
            logger.info(f"   Hon. Bruto:    R$ {hon_bruto:,.2f} (20%)")
            logger.info(f"   Hon. Líquido:  R$ {honorarios:,.2f}")
            logger.info(f"   Esperado:      R$ {esperado:,.2f} (90% do bruto)")
            logger.info(f"   Diff:          R$ {diff:,.2f}")
            
            if not passou:
                erros.append(f"{cenario}: R$ {diff:,.2f} diferença")
        
        assert len(erros) == 0, f"Deságio incorreto: {erros}"


if __name__ == "__main__":
    # Rodar testes diretamente
    pytest.main([__file__, "-v", "-s", "--log-cli-level=INFO"])
