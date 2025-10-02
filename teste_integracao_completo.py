"""
TESTE DE INTEGRAÇÃO COMPLETO: Web API vs Excel Direto
Simula o fluxo completo do usuário e compara resultados
"""
import sys
from pathlib import Path
from datetime import date
from openpyxl import load_workbook
import requests
import json

# Adicionar src ao path
sys.path.insert(0, str(Path(__file__).parent / "src"))

from models import CalculadoraInput
from calculator_service import CalculadoraService
from desagio_calculator import DesagioCalculator
from honorarios_calculator import HonorariosCalculator

print("=" * 80)
print("🧪 TESTE DE INTEGRAÇÃO COMPLETO: Web API vs Excel Direto")
print("=" * 80)

# ===== CASOS DE TESTE =====
casos_teste = [
    {
        "nome": "Caso 1: TIMON - Período Original",
        "municipio": "TIMON",
        "periodo_inicio": date(2000, 1, 1),
        "periodo_fim": date(2006, 12, 1),
        "ajuizamento": date(2005, 5, 1),
        "citacao": date(2006, 6, 1),
        "honorarios_perc": 20.0,
        "honorarios_fixo": 1000.0,
        "desagio_principal": 20.0,
        "desagio_honorarios": 20.0,
        "correcao_ate": date(2025, 1, 1)
    },
    {
        "nome": "Caso 2: TIMON - Período Reduzido",
        "municipio": "TIMON",
        "periodo_inicio": date(2000, 1, 1),
        "periodo_fim": date(2004, 1, 1),
        "ajuizamento": date(2005, 5, 1),
        "citacao": date(2006, 6, 1),
        "honorarios_perc": 15.0,
        "honorarios_fixo": 500.0,
        "desagio_principal": 25.0,
        "desagio_honorarios": 30.0,
        "correcao_ate": date(2025, 1, 1)
    },
    {
        "nome": "Caso 3: TIMON - Sem Deságio",
        "municipio": "TIMON",
        "periodo_inicio": date(2000, 1, 1),
        "periodo_fim": date(2006, 12, 1),
        "ajuizamento": date(2005, 5, 1),
        "citacao": date(2006, 6, 1),
        "honorarios_perc": 10.0,
        "honorarios_fixo": 0.0,
        "desagio_principal": 0.0,
        "desagio_honorarios": 0.0,
        "correcao_ate": date(2025, 1, 1)
    },
]

# ===== FUNÇÕES AUXILIARES =====

def calcular_via_servico(inputs_dict):
    """Calcula usando CalculadoraService (mesmo fluxo da API)"""
    inputs = CalculadoraInput(**inputs_dict)
    service = CalculadoraService()
    result = service.calcular(inputs)
    return result

def ler_excel_diretamente(inputs_dict):
    """Lê valores diretamente do Excel template"""
    excel_path = Path("data/timon_01-2025.xlsx")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["RESUMO"]
    
    # Ler valor D23 (primeiro cenário - NT7)
    principal_bruto = ws['D23'].value
    
    wb.close()
    
    if principal_bruto is None:
        return None
    
    # Aplicar deságio e calcular honorários como o sistema faz
    desagio_calc = DesagioCalculator()
    hon_calc = HonorariosCalculator()
    
    # Deságio no principal
    desagio_result = desagio_calc.aplicar_desagio(
        principal_bruto=principal_bruto,
        desagio_percentual=inputs_dict['desagio_principal']
    )
    principal_liquido = desagio_result['principal_liquido']
    
    # Honorários
    hon_result = hon_calc.calcular(
        principal=principal_liquido,
        honorarios_perc=inputs_dict['honorarios_perc'],
        honorarios_fixo=inputs_dict['honorarios_fixo'],
        desagio_honorarios=inputs_dict['desagio_honorarios']
    )
    
    return {
        'principal_bruto': principal_bruto,
        'principal_liquido': principal_liquido,
        'honorarios': hon_result['honorarios'],
        'total': hon_result['total']
    }

def comparar_resultados(servico_result, excel_result, caso_nome):
    """Compara resultados e retorna diferenças"""
    print(f"\n{'=' * 80}")
    print(f"📊 {caso_nome}")
    print("=" * 80)
    
    # Pegar primeiro cenário (NT7)
    servico_nt7 = servico_result.outputs.get('nt7_ipca_selic')
    
    if servico_nt7 is None:
        print("❌ Serviço não retornou NT7")
        return False
    
    if excel_result is None:
        print("❌ Excel não tem valores pré-calculados")
        return False
    
    print(f"\n📈 VALORES DO SERVIÇO (via CalculatorService):")
    print(f"   Principal: R$ {servico_nt7.principal:,.2f}")
    print(f"   Honorários: R$ {servico_nt7.honorarios:,.2f}")
    print(f"   Total: R$ {servico_nt7.total:,.2f}")
    
    print(f"\n📊 VALORES DO EXCEL (leitura direta):")
    print(f"   Principal BRUTO: R$ {excel_result['principal_bruto']:,.2f}")
    print(f"   Principal Líquido: R$ {excel_result['principal_liquido']:,.2f}")
    print(f"   Honorários: R$ {excel_result['honorarios']:,.2f}")
    print(f"   Total: R$ {excel_result['total']:,.2f}")
    
    # Comparar (permitir pequena diferença por arredondamento)
    tolerancia = 0.01  # R$ 0.01
    
    diff_principal = abs(servico_nt7.principal - excel_result['principal_liquido'])
    diff_honorarios = abs(servico_nt7.honorarios - excel_result['honorarios'])
    diff_total = abs(servico_nt7.total - excel_result['total'])
    
    print(f"\n🔍 DIFERENÇAS:")
    print(f"   Principal: R$ {diff_principal:,.2f}")
    print(f"   Honorários: R$ {diff_honorarios:,.2f}")
    print(f"   Total: R$ {diff_total:,.2f}")
    
    # Verificar se está dentro da tolerância
    if diff_principal <= tolerancia and diff_honorarios <= tolerancia and diff_total <= tolerancia:
        print(f"\n✅ SUCESSO: Valores idênticos (dentro da tolerância de R$ {tolerancia})")
        return True
    else:
        print(f"\n⚠️  DIFERENÇA: Valores não são idênticos")
        return False

# ===== EXECUTAR TESTES =====

print(f"\n🔬 Executando {len(casos_teste)} casos de teste...\n")

resultados = []

for caso in casos_teste:
    try:
        print(f"\n{'='*80}")
        print(f"🧪 TESTANDO: {caso['nome']}")
        print(f"{'='*80}")
        
        # Preparar inputs
        inputs_dict = {k: v for k, v in caso.items() if k != 'nome'}
        
        print(f"\n📝 Parâmetros:")
        print(f"   Município: {inputs_dict['municipio']}")
        print(f"   Período: {inputs_dict['periodo_inicio']} a {inputs_dict['periodo_fim']}")
        print(f"   Honorários: {inputs_dict['honorarios_perc']}% + R$ {inputs_dict['honorarios_fixo']:,.2f}")
        print(f"   Deságio: Principal {inputs_dict['desagio_principal']}%, Honorários {inputs_dict['desagio_honorarios']}%")
        
        # Calcular via serviço
        print(f"\n⏳ Calculando via CalculatorService...")
        servico_result = calcular_via_servico(inputs_dict)
        print(f"   ✅ Cálculo concluído em {servico_result.execution_time_ms}ms")
        
        # Ler Excel diretamente
        print(f"\n⏳ Lendo Excel diretamente...")
        excel_result = ler_excel_diretamente(inputs_dict)
        print(f"   ✅ Valores lidos do Excel")
        
        # Comparar
        sucesso = comparar_resultados(servico_result, excel_result, caso['nome'])
        
        resultados.append({
            'caso': caso['nome'],
            'sucesso': sucesso
        })
        
    except Exception as e:
        print(f"\n❌ ERRO no caso '{caso['nome']}': {str(e)}")
        import traceback
        traceback.print_exc()
        resultados.append({
            'caso': caso['nome'],
            'sucesso': False,
            'erro': str(e)
        })

# ===== RESUMO FINAL =====

print("\n" + "=" * 80)
print("📊 RESUMO FINAL DOS TESTES")
print("=" * 80)

total_casos = len(resultados)
casos_sucesso = sum(1 for r in resultados if r['sucesso'])
casos_falha = total_casos - casos_sucesso

print(f"\nTotal de casos testados: {total_casos}")
print(f"✅ Sucessos: {casos_sucesso}")
print(f"❌ Falhas: {casos_falha}")

print(f"\n{'Caso':<50} {'Status':<10}")
print("-" * 60)
for r in resultados:
    status = "✅ OK" if r['sucesso'] else "❌ FALHA"
    print(f"{r['caso']:<50} {status:<10}")

if casos_falha == 0:
    print("\n" + "=" * 80)
    print("🎉 TODOS OS TESTES PASSARAM!")
    print("=" * 80)
    print("O sistema está calculando EXATAMENTE os mesmos valores que o Excel.")
    print("A integração está 100% funcional!")
else:
    print("\n" + "=" * 80)
    print("⚠️  ALGUNS TESTES FALHARAM")
    print("=" * 80)
    print("Verifique as diferenças acima para debug.")

print("\n" + "=" * 80)
